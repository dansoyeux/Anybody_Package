# -*- coding: utf-8 -*-
"""
Created on Mon Nov 13 11:03:46 2023

@author: Dan
"""

import pandas as pd
import numpy as np
from Anybody_Package.Anybody_LoadOutput.Tools import array_to_dictionary

from openpyxl.utils.cell import get_column_letter

# from Anybody_Package.Anybody_LoadOutput.Tools import array_to_dictionary

"""
to comment once on tools
"""


def seperate_data_frame_by_category(data_frame):
    """
    seperates a pandas data_frame by categories

    the first line of the entered dataframe has a category name or nan values.
    groups the data from the column where there is a category name to the column before the next category name.

    -----------------------------------------------------------------------
    returns
    seperated_categories_dictionary :  dictionay where each key is the name of that category


    ex : col numbers :    1       2    3    4    5       6    7      8      9    10
         line 1 :       "cat_1", NaN, NaN, NaN "cat_2", NaN, NaN, "cat_3", NaN, NaN

    ---------------------------------------------------
    return

    seperated_categories_dictionary : dict : contains on each entry the data_frame of a category

    seperated_categories_dictionary["cat_1"] : column 1 to 4
    seperated_categories_dictionary["cat_2"] : column 5 to 7
    seperated_categories_dictionary["cat_3"] : column 8 to 10
    """
    # Selects the line that stores the category names
    category_line = data_frame.iloc[0, :]

    # gets the indexes of each column
    category_line_index = category_line.index.to_numpy()

    # list containing the index of the columns that have a category name (not NaN)
    category_first_value = category_line_index[category_line.notna().to_numpy()]

    # list of all categories
    category_name_list = category_line[category_line.notna().to_numpy()].tolist()

    seperated_categories_dictionary = {}

    # Gets the data for each category
    for category_index, category_name in enumerate(category_name_list):
        # select the data for each author
        # for every authors except the last
        if not category_name == category_name_list[-1]:
            current_category_data = data_frame.loc[:, category_first_value[category_index]: category_first_value[category_index + 1] - 1]
        # for the last author
        else:
            current_category_data = data_frame.loc[:, category_first_value[category_index]: category_line_index[-1]]

        # deletes the row that has author names
        current_category_data = current_category_data.iloc[1: len(current_category_data), :]

        seperated_categories_dictionary[category_name] = current_category_data

    return seperated_categories_dictionary


def variable_data_frame_to_dictionary(ExcelFile, variable_dataframe, variable_informations, author_name, muscle_name=""):
    """
    function that takes the variables from the cariable_data and returns them as a dictionary containing its description, component sequence and values by component

    variable_dataframe : line 0 : Variable names
                    line 1 : component names
                    lines 2:end : values

    variable_informations : dict : informations about the x and y variables (component sequence, multiply factors, descriptions)
                          : variable_informations = {variable_x_name: {"VariableDescription": variable_x_description, "MultiplyFactor": variable_x_multiply_factor},
                                                     variable_y_name: {"VariableDescription": variable_y_description, "MultiplyFactor": variable_y_multiply_factor}
                                                     }

    -------------------------------------------------
    return
    loaded_variables : dict : informations about the loaded variables

    variable_x_dictionary : dict : Contains the values, description and component sequence of the x variable
    variable_y_dictionary : dict : Contains the values, description and component sequence of the y variable
    nStep : int the number of steps in the data

    """

    # deletes lines full of nan values
    variable_dataframe = variable_dataframe.dropna(axis=0, how='all')

    # gets the names of the variables
    variable_names_list = list(variable_informations.keys())
    variable_x_name = variable_names_list[0]
    variable_y_name = variable_names_list[1]

    # line with variables
    variable_line = variable_dataframe.iloc[0, :]
    # line with components
    component_line = variable_dataframe.iloc[1, :]

    # index of the line where there is the first value
    first_value_index = variable_dataframe.index.values[2]
    last_value_index = variable_dataframe.index.values[-1]

    # column index where the x_variables and y_variables are stored
    variable_x_index = variable_line.iloc[0::2].index.values
    variable_y_index = variable_line.iloc[1::2].index.values


    # names entered as the x and y variables
    variable_x_columns = variable_line.iloc[0::2]
    variable_y_columns = variable_line.iloc[1::2]

    # Checks that all the variables names entered on the variable data column match the name of the variable name
    variable_x_name_error_index = variable_x_columns[variable_x_columns != variable_x_name].index.values
    variable_y_name_error_index = variable_y_columns[variable_y_columns != variable_y_name].index.values
    # variable_y_name_error_index = variable_line[(variable_line.loc[variable_x_index] != variable_x_name) & (variable_line != variable_y_name)].index.values

    # if error on x
    if not variable_x_name_error_index.size == 0:
        # gets the column name where the errors are
        column_names_x_error = [get_column_letter(error_index + 1) for error_index in variable_x_name_error_index]
        ExcelFile.close()
        if muscle_name:
            raise ValueError(f"For the author '{author_name}' in the muscle '{muscle_name}', the columns {column_names_x_error} must be equal to the variable x name = '{variable_x_name}'")
        else:
            raise ValueError(f"For the author '{author_name}', the columns {column_names_x_error} must be equal to the variable x name = '{variable_x_name}'")

    # if error on y
    if not variable_y_name_error_index.size == 0:
        # gets the column name where the errors are
        column_names_y_error = [get_column_letter(error_index + 1) for error_index in variable_y_name_error_index]
        ExcelFile.close()
        if muscle_name:
            raise ValueError(f"For the author '{author_name}' in the muscle '{muscle_name}', the columns {column_names_y_error} must be equal to the variable y name = '{variable_y_name}'")
        else:
            raise ValueError(f"For the author '{author_name}', the columns {column_names_y_error} must be equal to the variable y name = '{variable_y_name}'")

    # variables component sequence
    variable_x_component_sequence = component_line.loc[variable_x_index].tolist()
    variable_y_component_sequence = component_line.loc[variable_y_index].tolist()

    # number of component for the y variable
    n_composantes = len(variable_y_index)

    # gets all the x and y values
    variable_x_array = variable_dataframe.loc[first_value_index: last_value_index, variable_x_index].to_numpy(dtype=float)
    variable_y_array = variable_dataframe.loc[first_value_index: last_value_index, variable_y_index].to_numpy(dtype=float)

    # sets the dictionary that will store all the variable informations
    loaded_variables = variable_informations

    loaded_variables[variable_x_name]["SequenceComposantes"] = variable_y_component_sequence

    # Adds the component sequence to the loaded variables informations
    loaded_variables[variable_y_name]["SequenceComposantes"] = variable_y_component_sequence

    # all x component names should match
    if not np.all(np.array(variable_x_component_sequence) == np.array(variable_x_component_sequence)[0]):
        ExcelFile.close()
        raise ValueError(f"For the variable : '{variable_y_name}', for the author : '{author_name}'\nEach x variable component must be the same")

    # For each x variable there must be a y variable associated
    if not len(variable_x_index) == len(variable_y_index):
        ExcelFile.close()
        raise ValueError(f"For the variable : '{variable_y_name}', for the author : '{author_name}'\neach x variable must be associated to a y variable")

    # if there is a nan in the variable data, it means that the component have different lengths
    # Deletes these nan values
    nan_value_in_dataframe = bool(variable_dataframe.isnull().values.any())

    # if there are only one component for y, x and y values should have the same length
    if n_composantes == 1 and nan_value_in_dataframe:
        ExcelFile.close()
        raise ValueError(f"For the variable : '{variable_y_name}', for the author : '{author_name}'\nthe x and y values must have the same length")

    # Transforms the obtained arrays to a dictionary and doesn't calculate the total
    variable_x_dictionary = array_to_dictionary(variable_x_array, **loaded_variables[variable_x_name], total_on=False)
    variable_y_dictionary = array_to_dictionary(variable_y_array, **loaded_variables[variable_y_name], total_on=False)

    return variable_x_dictionary, variable_y_dictionary, loaded_variables


def interpolate_y_variable(variable_x_array, variable_y_array, x_variable_interpolation_values):
    """
    Finds an interpolation function y = f(x) and finds y : y_variable_interpolate = f(x_variable_interpolation_values)

    The interpolation function is a CubicSpline

    ------------------------------------------------------------
    return
    variable_y_interpolate

    """
    from scipy.interpolate import CubicSpline

    # Interpolates the y variable using min and max x variable entered so that every y component correspond to the same x values
    Interpolation_Function = CubicSpline(variable_x_array, variable_y_array)

    # gets the interpolated y value
    variable_y_interpolated = Interpolation_Function(x_variable_interpolation_values)

    return variable_y_interpolated


def get_excel_sheet_variable_informations(variable_informations_data):
    """
    function that puts in a dictionary every information about the x and y variables from the current sheet variable informations data

    variable descriptions and multiply factors

    ----------------------------------------------------------
    return
            variable_informations : dict : informations about the x and y variables (component sequence, multiply factors, descriptions)
                                  : variable_informations = {variable_x_name: {"VariableDescription": variable_x_description, "MultiplyFactor": variable_x_multiply_factor},
                                                             variable_y_name: {"VariableDescription": variable_y_description, "MultiplyFactor": variable_y_multiply_factor}
                                                             }

    """
    # gets the index of the variables informations
    x_variable_index = int(variable_informations_data[variable_informations_data.loc[:, 0] == "Variable x"].index.values[0])
    y_variable_index = int(variable_informations_data[variable_informations_data.loc[:, 0] == "Variable y"].index.values[0])

    # gets the data frame for each variable informations
    x_variable_informations = variable_informations_data.loc[x_variable_index: x_variable_index + 2, 1]
    y_variable_informations = variable_informations_data.loc[y_variable_index: y_variable_index + 2, 1].dropna()

    # variable x informations
    variable_x_name = x_variable_informations.iloc[0]
    variable_x_description = x_variable_informations.iloc[1]
    variable_x_multiply_factor = x_variable_informations.iloc[2]

    # variable_y informations
    variable_y_name = y_variable_informations.iloc[0]
    variable_y_description = y_variable_informations.iloc[1]
    variable_y_multiply_factor = y_variable_informations.iloc[2]

    variable_informations = {variable_x_name: {"VariableDescription": variable_x_description, "MultiplyFactor": variable_x_multiply_factor},
                             variable_y_name: {"VariableDescription": variable_y_description, "MultiplyFactor": variable_y_multiply_factor}
                             }

    return variable_informations, variable_x_name, variable_y_name


def get_Excel_sheet_variables(ExcelFile, current_sheet_name):
    """
    loads the variable from the selected excel sheet
    """
    sheet_data = pd.read_excel(ExcelFile, current_sheet_name, header=None)

    # gets the variables informations
    variable_informations_data = sheet_data.iloc[:, 0: 2]

    variable_informations, variable_x_name, variable_y_name = get_excel_sheet_variable_informations(variable_informations_data)

    # gets the variable data
    variables_data = sheet_data.iloc[:, 2:len(sheet_data.iloc[0, :]) + 1]

    result_dictionary = {}

    # seperates the dataframe by author names
    author_data = seperate_data_frame_by_category(variables_data)

    # goes through
    for author_name, current_author_data in author_data.items():

        # in case of a muscle variable, seperates the sheet_data by muscle name
        if "Muscle" in current_sheet_name:
            # seperates the data by variables
            muscle_author_data = seperate_data_frame_by_category(current_author_data)

            # initializes the result dictionary with a "Muscles" directory to store muscle variables
            result_dictionary[author_name] = {"Muscles": {}}

            # goes through each muscle data for the current author
            for muscle_name, current_muscle_data in muscle_author_data.items():

                variable_x_dictionary, variable_y_dictionary, loaded_variables = variable_data_frame_to_dictionary(ExcelFile, current_muscle_data, variable_informations, author_name, muscle_name)

                # stores the current muscle y_variable
                result_dictionary[author_name]["Muscles"][muscle_name] = {muscle_name: {variable_y_name: variable_y_dictionary}}

                # stores the x variable as a muscle variable
                result_dictionary[author_name]["Muscles"][muscle_name][muscle_name][variable_x_name] = variable_x_dictionary

            # stores the informations about the loaded variables
            result_dictionary[author_name]["Loaded Variables"] = {"Variables": {variable_x_name: loaded_variables[variable_x_name]}, "MuscleVariables": {variable_y_name: loaded_variables[variable_y_name]}}

        # if the loaded variable is a normal variable
        else:
            # transforms the variables data into a result dictionary
            variable_x_dictionary, variable_y_dictionary, loaded_variables = variable_data_frame_to_dictionary(ExcelFile, current_author_data, variable_informations, author_name)

            # for a normal variable
            result_dictionary[author_name] = {variable_x_name: variable_x_dictionary, variable_y_name: variable_y_dictionary}

            # stores the informations about the loaded variables
            result_dictionary[author_name]["Loaded Variables"] = {"Variables": loaded_variables}

        result_dictionary[author_name]["Loaded Variables"]["Data Source"] = "Literature"

    return result_dictionary


def load_literature_data(file_name, directory_path=""):
    """
    Loads literature data from the excel template

    file_name : str : name of the excel file

    directory_path : str : path of the directory containing the file from the current directory

    """

    extension = "xlsx"

    # builds the path of the excel file to load
    if directory_path:
        file_path = f"{directory_path}/{file_name}.{extension}"
    else:
        file_path = f"{file_name}.{extension}"

    ExcelFile = pd.ExcelFile(file_path)

    # gets the list from the excel file
    sheet_names_list = ExcelFile.sheet_names

    # deletes the names of the template sheets
    sheet_names_list = [sheet_name for sheet_name in sheet_names_list if "Template" not in sheet_name]

    result_dictionary = {}

    # goes through every excel sheet and gets its data
    for current_sheet_name in sheet_names_list:

        # gets the data for the current y variable
        sheet_result_dictionary = get_Excel_sheet_variables(ExcelFile, current_sheet_name)

        # gets the name of the sheet (without the Muscle. at the beginning)
        variable_y_name = current_sheet_name.split(".")[-1]

        # stores each sheet data as an entry in the result dictionary with the y variable name
        result_dictionary[variable_y_name] = sheet_result_dictionary

    ExcelFile.close()

    return result_dictionary
